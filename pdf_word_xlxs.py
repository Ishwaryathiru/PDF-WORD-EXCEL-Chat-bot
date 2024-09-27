import openpyxl
import streamlit as st
from PyPDF2 import PdfReader
from langchain.text_splitter import RecursiveCharacterTextSplitter
import os
from langchain_google_genai import GoogleGenerativeAIEmbeddings
import google.generativeai as genai
from langchain_community.vectorstores import FAISS
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate
from dotenv import load_dotenv

load_dotenv()
genai.configure(api_key=os.getenv('GOOGLE_API_KEY')) 

def get_pdf_text(path):
    text = ""
    for pdf in path:
        pr = PdfReader(pdf)
        for pg in pr.pages:
            text += pg.extract_text()
    return text


def get_text_chunks(text):
    ts = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
    chunks = ts.split_text(text)
    return chunks


def get_vector_store(text_chunks):
    emb = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    vs = FAISS.from_texts(text_chunks, embedding=emb)
    #vs.save_local(f'C:/Users/ishwarya.thirumuruga/Desktop/Ishwarya/Pdf chat/uploaded')
    vs.save_local(f'vectors_storage')


def get_conversational_chain():
    prmt_temp = """
    You are a instructor who reads the context and answer the question accurately.
    Answer the question as detailed as possible from the provided context. 
    If the answer is not found in the context, just say "Unable to find the answer", don't provide a wrong answer.
    context:\n{context}\n
    Question:\n{question}\n
    Answer:
    """
    model = ChatGoogleGenerativeAI(model="gemini-pro", temperature=0.3)
    prompt = PromptTemplate(template=prmt_temp, input_variables=["context", "question"])
    chain = load_qa_chain(model, chain_type="stuff", prompt=prompt)
    return chain


def user_input(ques):
    emb = GoogleGenerativeAIEmbeddings(model='models/embedding-001')
    #new_db = FAISS.load_local(f'C:/Users/ishwarya.thirumuruga/Desktop/Ishwarya/Pdf chat/uploaded', emb, allow_dangerous_deserialization=True)
    new_db = FAISS.load_local(f'vectors_storage', emb, allow_dangerous_deserialization=True)
    dox = new_db.similarity_search(ques)
    chain = get_conversational_chain()
    res = chain.invoke(
        {"input_documents": dox, "question": ques},
    )
    return res["output_text"]


import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import PyPDF2
import docx
import docx2pdf
import builtins
from docx import Document
from openpyxl import workbook


def get_docx_text(path):
    text = ""
    doc = docx.Document(path)
    for para in doc.paragraphs:
        text += para.text
    return text


def get_excel_text(path):
    text = ""
    wb = openpyxl.load_workbook(path)
    for sheet in wb:
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]  # Get headers from first row
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from second row (data rows)
            text += '\n'.join(f"{header}: {cell}" for header, cell in zip(headers, row)) + '\n\n'  # Concatenate column name with value 
    return text


mas_read=""
fname=[]
def open_pdf():
    global mas_read 
    global fname
    file_paths = filedialog.askopenfilenames(title="Select a PDF", filetypes=[("PDF Files", "*.pdf"),("Word Documents","*.docx"),("Excel Files", "*.xlsx")])
    for file_path in file_paths:
        if file_path:
            if file_path.endswith('.pdf'):
                pdf_content=get_pdf_text([file_path])
                mas_read+=' '+pdf_content
                chunks = get_text_chunks(mas_read)
                get_vector_store(chunks)
                fname.append(file_path.split('/')[-1])
                #print(fname)
            elif file_path.endswith('.docx'):
                docx_content = get_docx_text(file_path)
                mas_read += ' '+docx_content
                chunks = get_text_chunks(mas_read)
                get_vector_store(chunks)
                fname.append(file_path.split('/')[-1])
                #print(fname)
            elif file_path.endswith('.xlsx'):
                excel_content = get_excel_text(file_path)
                mas_read += ' '+excel_content
                chunks = get_text_chunks(mas_read)
                get_vector_store(chunks)
                fname.append(file_path.split('/')[-1])
                print(fname)
    messagebox.showinfo("Upload Status", "File uploaded successfully!")
    # print(mas_read)


import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox, simpledialog, scrolledtext

def view_uploaded_files():
    if fname:
        messagebox.showinfo("Uploaded Files", "\n".join(fname))
    else:
        messagebox.showinfo("Uploaded Files", "No files uploaded yet.")

def submit_question_answer():
    question = question_entry.get()
    answer = user_input(question)
    display_chat_bubble(question, "right")
    display_chat_bubble(answer, "left")
    question_entry.delete(0, tk.END)
    chat_area.yview(tk.END)

def display_chat_bubble(text, side):
    bubble_frame = tk.Frame(chat_area, bg="#ffffff", pady=5)
    
    bubble = tk.Label(
        bubble_frame, text=text, wraplength=350, padx=10, pady=10,
        bg="#dcf8c6" if side == "left" else "#ffffff", fg="black", font=("Helvetica", 12),
        bd=0, relief=tk.SOLID, borderwidth=1
    )
    
    bubble_frame.grid_rowconfigure(0, weight=1)
    bubble_frame.grid_columnconfigure(0, weight=1)
    
    if side == "right":
        bubble_frame.grid(row=0, column=1, sticky="e", padx=10, pady=5)
        bubble.grid(row=0, column=1, sticky="e")
    else:
        bubble_frame.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        bubble.grid(row=0, column=0, sticky="w")

    chat_area.window_create(tk.END, window=bubble_frame)
    chat_area.insert(tk.END, "\n")

# Create the main window
root = tk.Tk()
root.title("Chatbot PDF Viewer and Q&A")
root.geometry("430x650")
root.config(bg="#e5ddd5")

# Display area for chat history
chat_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, state=tk.DISABLED, font=("Helvetica", 12), bg="#ffffff", fg="black", bd=0)
chat_area.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

button_frame = tk.Frame(root, bg="#e5ddd5")
button_frame.pack(side=tk.BOTTOM, pady=10)

# Upload button at the bottom
upload_button = tk.Button(button_frame, text="Upload Files", command=open_pdf, bg="#34b7f1", fg="white", font=("Helvetica", 10, "bold"))
upload_button.grid(row=0, column=0, padx=10)

# View Uploads button next to Upload Files
view_uploads_button = tk.Button(button_frame, text="View Uploads", command=view_uploaded_files, bg="#34b7f1", fg="white", font=("Helvetica", 10, "bold"))
view_uploads_button.grid(row=0, column=1, padx=10)

# Label for uploaded files
uploaded_label = tk.Label(root, text="", bg="#e5ddd5", fg="black", font=("Helvetica", 10))
uploaded_label.pack()

# Question entry area
question_entry = tk.Entry(root, font=("Helvetica", 12), width=50, bd=1, relief=tk.SOLID)
question_entry.pack(padx=10, pady=10, fill=tk.X)

# Submit button for the question
submit_button = tk.Button(root, text="Send", command=submit_question_answer, bg="#25d366", fg="white", font=("Helvetica", 10, "bold"))
submit_button.pack(padx=10, pady=10)

# Run the Tkinter main loop
root.mainloop()